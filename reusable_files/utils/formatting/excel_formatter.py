"""
ExcelFormatter - Aplicación de estilos a archivos Excel
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional


class ExcelFormatter:
    """Aplica estilos y formato a archivos Excel"""
    
    @staticmethod
    def apply_header_style(worksheet, header_row: int = 1):
        """
        Aplica estilo a encabezados
        
        Args:
            worksheet: Hoja de Excel
            header_row: Número de fila de encabezados
        """
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[header_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    @staticmethod
    def apply_freeze_panes(worksheet, row: int = 1, col: int = 0):
        """
        Congela paneles
        
        Args:
            worksheet: Hoja de Excel
            row: Fila desde donde congelar
            col: Columna desde donde congelar
        """
        cell = worksheet.cell(row + 1, col + 1)
        worksheet.freeze_panes = cell.coordinate
    
    @staticmethod
    def apply_autofilter(worksheet, max_row: int, max_col: int):
        """
        Aplica autofiltro
        
        Args:
            worksheet: Hoja de Excel
            max_row: Última fila con datos
            max_col: Última columna con datos
        """
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    @staticmethod
    def auto_adjust_columns(worksheet, min_width: int = 10, max_width: int = 50):
        """
        Ajusta ancho de columnas automáticamente
        
        Args:
            worksheet: Hoja de Excel
            min_width: Ancho mínimo
            max_width: Ancho máximo
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def apply_alternating_rows(worksheet, start_row: int = 2):
        """
        Aplica color alternado a filas
        
        Args:
            worksheet: Hoja de Excel
            start_row: Fila inicial
        """
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            if idx % 2 == 0:
                for cell in row:
                    if cell.value:
                        cell.fill = light_fill
    
    @staticmethod
    def apply_borders(worksheet):
        """
        Aplica bordes a todas las celdas con contenido
        
        Args:
            worksheet: Hoja de Excel
        """
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    @staticmethod
    def format_excel_complete(
        worksheet,
        apply_header: bool = True,
        apply_freeze: bool = True,
        apply_filter: bool = True,
        apply_adjust: bool = True,
        apply_alternating: bool = False,
        apply_borders: bool = False
    ):
        """
        Aplica formato completo a una hoja
        
        Args:
            worksheet: Hoja de Excel
            apply_header: Aplicar estilo a encabezados
            apply_freeze: Congelar paneles
            apply_filter: Aplicar autofiltro
            apply_adjust: Ajustar anchos de columna
            apply_alternating: Aplicar colores alternados
            apply_borders: Aplicar bordes
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        if apply_header:
            ExcelFormatter.apply_header_style(worksheet)
        
        if apply_freeze:
            ExcelFormatter.apply_freeze_panes(worksheet)
        
        if apply_filter:
            ExcelFormatter.apply_autofilter(worksheet, max_row, max_col)
        
        if apply_adjust:
            ExcelFormatter.auto_adjust_columns(worksheet)
        
        if apply_alternating:
            ExcelFormatter.apply_alternating_rows(worksheet)
        
        if apply_borders:
            ExcelFormatter.apply_borders(worksheet)