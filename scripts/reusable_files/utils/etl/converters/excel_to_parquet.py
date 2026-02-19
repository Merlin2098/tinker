from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def normalize_token(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", normalize_text(value))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized.upper()


class ExcelToParquetConverter:
    """Read selected sheet/columns from Excel and persist parquet outputs."""

    @staticmethod
    def read_sheet_by_config(
        excel_path: str | Path,
        *,
        sheet_name: str,
        json_config: dict[str, Any],
        read_all_as_string: bool = True,
    ) -> tuple[pl.DataFrame, str]:
        header_row = int(json_config.get("header_row", 1))
        data_start_row = int(json_config.get("data_start_row", header_row + 1))
        keep_columns = json_config.get("keep_columns") or json_config.get("columns_to_retain")
        anchor_column = json_config.get("anchor_column")

        if not keep_columns:
            raise ValueError("json_config must define keep_columns or columns_to_retain")

        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        normalized_sheets = {normalize_token(name): name for name in workbook.sheetnames}
        expected_token = normalize_token(sheet_name)
        if expected_token not in normalized_sheets:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}. Available: {available}")

        actual_sheet = normalized_sheets[expected_token]
        worksheet = workbook[actual_sheet]

        row_iter = worksheet.iter_rows(min_row=header_row, values_only=True)
        header_values = next(row_iter, None)
        if header_values is None:
            raise ValueError(
                f"Sheet '{actual_sheet}' has no rows at or below configured header_row={header_row}"
            )

        headers = [normalize_text(value) for value in header_values]
        token_to_column = {
            normalize_token(column): column
            for column in headers
            if column not in ("", None)
        }

        selected_columns: list[str] = []
        missing_columns: list[str] = []
        for configured in keep_columns:
            match = token_to_column.get(normalize_token(configured))
            if match is None:
                missing_columns.append(configured)
            else:
                selected_columns.append(match)

        if missing_columns:
            raise ValueError(f"Missing configured columns in sheet '{actual_sheet}': {missing_columns}")

        def clean_value(value: Any) -> Any:
            if value is None:
                return None
            if read_all_as_string:
                text = normalize_text(value)
                if text in {"", "nan", "None"}:
                    return None
                return text
            return value

        anchor_real = None
        if anchor_column:
            anchor_real = token_to_column.get(normalize_token(anchor_column))
            if anchor_real is None:
                raise ValueError(f"Anchor column '{anchor_column}' not found in sheet '{actual_sheet}'")

        index_lookup = {header: idx for idx, header in enumerate(headers)}
        payload_rows: list[dict[str, Any]] = []
        for offset, raw_row in enumerate(row_iter, start=header_row + 1):
            if offset < data_start_row:
                continue

            record: dict[str, Any] = {}
            for column in selected_columns:
                col_idx = index_lookup[column]
                value = raw_row[col_idx] if col_idx < len(raw_row) else None
                record[column] = clean_value(value)

            if all(record[column] is None for column in selected_columns):
                continue

            if anchor_real is not None:
                anchor_idx = index_lookup[anchor_real]
                anchor_value = raw_row[anchor_idx] if anchor_idx < len(raw_row) else None
                if clean_value(anchor_value) is None:
                    continue

            record["_source_row"] = offset
            record["_source_file"] = Path(excel_path).name
            record["_source_sheet"] = actual_sheet
            payload_rows.append(record)

        output_columns = selected_columns + ["_source_row", "_source_file", "_source_sheet"]
        if not payload_rows:
            empty_payload = pl.DataFrame({column: [] for column in output_columns})
            return empty_payload, actual_sheet

        payload = pl.DataFrame(payload_rows, infer_schema_length=2000).select(output_columns)
        return payload, actual_sheet

    @staticmethod
    def to_parquet(frame: object, output_path: str | Path) -> None:
        target = Path(output_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        if isinstance(frame, pl.DataFrame):
            # Keep explicit column sequence when persisting parquet.
            frame.select(frame.columns).write_parquet(target)
            return
        if hasattr(frame, "to_parquet"):
            # Pandas compatibility path; enforce current column order explicitly.
            if hasattr(frame, "columns") and hasattr(frame, "__getitem__"):
                ordered = frame[list(frame.columns)]
            else:
                ordered = frame
            ordered.to_parquet(target, index=False)
            return
        raise TypeError(f"Unsupported frame type for parquet export: {type(frame)!r}")
