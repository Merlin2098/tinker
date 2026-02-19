from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Excel worksheet formatting helpers."""

    @staticmethod
    def apply_header_style(worksheet, header_row: int = 1) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[header_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

    @staticmethod
    def apply_freeze_panes(worksheet, row: int = 1, col: int = 0) -> None:
        worksheet.freeze_panes = worksheet.cell(row + 1, col + 1).coordinate

    @staticmethod
    def apply_autofilter(worksheet, max_row: int, max_col: int) -> None:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    @staticmethod
    def auto_adjust_columns(worksheet, min_width: int = 10, max_width: int = 50) -> None:
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)

    @staticmethod
    def apply_alternating_rows(worksheet, start_row: int = 2) -> None:
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
            if row_idx % 2 == 0:
                for cell in row:
                    if cell.value:
                        cell.fill = light_fill

    @staticmethod
    def apply_borders(worksheet) -> None:
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = border
