#!/usr/bin/env python3
"""
Canonical execution wrapper for the `excel_explorer` skill.
"""

from __future__ import annotations

from typing import Any

try:
    from agents.tools.wrappers._explorer_common import (
        file_stats,
        parse_bool,
        parse_int,
        resolve_repo_path,
        to_jsonable,
    )
except ImportError:
    from wrappers._explorer_common import file_stats, parse_bool, parse_int, resolve_repo_path, to_jsonable


def _sheet_name(value: Any) -> str | int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip():
        stripped = value.strip()
        if stripped.isdigit():
            return int(stripped)
        return stripped
    raise ValueError("sheet_name must be a non-empty string or integer when provided.")


def run(args: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(args, dict):
        raise ValueError("Wrapper args must be a JSON object.")

    resolved_str, resolved = resolve_repo_path(
        args.get("path"),
        field_name="path",
        allowed_suffixes={".xlsx", ".xlsm", ".xltx", ".xltm"},
    )
    data_only = parse_bool(args.get("data_only"), field="data_only", default=True)
    read_only = parse_bool(args.get("read_only"), field="read_only", default=True)
    preview_rows = parse_int(args.get("preview_rows"), field="preview_rows", default=5, minimum=0, maximum=100)
    selected = _sheet_name(args.get("sheet_name"))

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from exc

    workbook = openpyxl.load_workbook(resolved, read_only=read_only, data_only=data_only)
    try:
        sheet_summaries: list[dict[str, Any]] = []
        for ws in workbook.worksheets:
            sheet_summaries.append(
                {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                }
            )

        if selected is None:
            worksheet = workbook.worksheets[0]
        elif isinstance(selected, int):
            if selected < 0 or selected >= len(workbook.worksheets):
                raise ValueError(f"sheet_name index out of range: {selected}")
            worksheet = workbook.worksheets[selected]
        else:
            if selected not in workbook.sheetnames:
                raise ValueError(f"sheet_name not found: {selected}")
            worksheet = workbook[selected]

        row_iter = worksheet.iter_rows(values_only=True)
        header_raw = next(row_iter, None)
        headers: list[str] = []
        if header_raw is not None:
            headers = [str(v).strip() if v is not None else f"column_{idx + 1}" for idx, v in enumerate(header_raw)]

        preview: list[dict[str, Any]] = []
        for row in row_iter:
            if len(preview) >= preview_rows:
                break
            row_values = list(row)
            row_values += [None] * max(0, len(headers) - len(row_values))
            preview.append({headers[idx]: to_jsonable(row_values[idx]) for idx in range(len(headers))})
    finally:
        workbook.close()

    return {
        "status": "ok",
        "skill": "excel_explorer",
        "path": args.get("path"),
        "resolved_path": resolved_str,
        "read_only": read_only,
        "data_only": data_only,
        "sheet_count": len(sheet_summaries),
        "sheet_summaries": sheet_summaries,
        "selected_sheet": worksheet.title,
        "columns": headers,
        "column_count": len(headers),
        "rows_preview": preview,
        **file_stats(resolved),
    }


