#!/usr/bin/env python3
"""
Canonical execution wrapper for the `read_excel_polars_openpyxl` skill.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any


def _project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _resolve_excel_path(raw_path: Any) -> tuple[str, Path]:
    if not isinstance(raw_path, str) or not raw_path.strip():
        raise ValueError("path is required and must be a non-empty string.")

    project_root = _project_root()
    candidate = Path(raw_path.strip())
    if not candidate.is_absolute():
        candidate = project_root / candidate
    resolved = candidate.resolve()

    try:
        resolved.relative_to(project_root)
    except ValueError as exc:
        raise ValueError(f"path must resolve inside project root: {resolved}") from exc

    if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("path must point to a .xlsx or .xlsm file.")
    if not resolved.exists():
        raise FileNotFoundError(f"Excel file not found: {resolved}")
    if not resolved.is_file():
        raise ValueError(f"path must be a file: {resolved}")

    return str(resolved), resolved


def _sheet_name(value: Any, default: int = 0) -> str | int:
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            raise ValueError("sheet_name cannot be empty when provided.")
        if stripped.isdigit():
            return int(stripped)
        return stripped
    raise ValueError("sheet_name must be a string or integer when provided.")


def _bool(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "y"}:
            return True
        if lowered in {"false", "0", "no", "n"}:
            return False
    raise ValueError(f"Invalid boolean value: {value!r}")


def _preview_rows(value: Any, default: int = 5) -> int:
    if value is None:
        return default
    if isinstance(value, int):
        parsed = value
    elif isinstance(value, str) and value.strip().isdigit():
        parsed = int(value.strip())
    else:
        raise ValueError("preview_rows must be an integer when provided.")
    if parsed < 0 or parsed > 100:
        raise ValueError("preview_rows must be between 0 and 100.")
    return parsed


def _use_columns(value: Any) -> list[str] | None:
    if value is None:
        return None
    if isinstance(value, list):
        if not value:
            raise ValueError("use_columns list cannot be empty.")
        columns = []
        for column in value:
            if not isinstance(column, str) or not column.strip():
                raise ValueError("use_columns items must be non-empty strings.")
            columns.append(column.strip())
        return columns
    raise ValueError("use_columns must be a list of strings when provided.")


def _normalize_headers(header_values: tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    normalized: list[str] = []
    for idx, raw in enumerate(header_values):
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = f"column_{idx + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        normalized.append(base if count == 0 else f"{base}_{count + 1}")
    return normalized


def _to_jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def run(args: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(args, dict):
        raise ValueError("Wrapper args must be a JSON object.")

    resolved_str, resolved = _resolve_excel_path(args.get("path"))
    requested_sheet = _sheet_name(args.get("sheet_name"), default=0)
    read_only = _bool(args.get("read_only"), default=True)
    data_only = _bool(args.get("data_only"), default=True)
    preview_rows = _preview_rows(args.get("preview_rows"), default=5)
    use_columns = _use_columns(args.get("use_columns"))

    try:
        import openpyxl
        import polars as pl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl and polars are required for read_excel_polars_openpyxl wrapper. "
            "Install with: pip install openpyxl polars"
        ) from exc

    workbook = openpyxl.load_workbook(resolved, read_only=read_only, data_only=data_only)
    try:
        if isinstance(requested_sheet, int):
            if requested_sheet < 0 or requested_sheet >= len(workbook.worksheets):
                raise ValueError(
                    f"sheet_name index out of range: {requested_sheet}. "
                    f"Available indices: 0..{len(workbook.worksheets) - 1}"
                )
            worksheet = workbook.worksheets[requested_sheet]
        else:
            if requested_sheet not in workbook.sheetnames:
                raise ValueError(
                    f"sheet_name not found: {requested_sheet}. "
                    f"Available sheets: {workbook.sheetnames}"
                )
            worksheet = workbook[requested_sheet]

        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise ValueError("Worksheet is empty. No header row found.")

        headers = _normalize_headers(tuple(header_row))
        rows: list[list[Any]] = []
        for row in row_iter:
            values = list(row)
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            elif len(values) > len(headers):
                values = values[: len(headers)]
            rows.append(values)
    finally:
        workbook.close()

    dataframe = pl.DataFrame(rows, schema=headers, orient="row")

    if use_columns:
        missing = [column for column in use_columns if column not in dataframe.columns]
        if missing:
            raise ValueError(f"use_columns contains unknown columns: {missing}")
        dataframe = dataframe.select(use_columns)

    rows_preview = [
        {key: _to_jsonable(value) for key, value in record.items()}
        for record in dataframe.head(preview_rows).to_dicts()
    ]

    return {
        "status": "ok",
        "skill": "read_excel_polars_openpyxl",
        "path": args.get("path"),
        "resolved_path": resolved_str,
        "sheet_name": worksheet.title,
        "read_only": read_only,
        "data_only": data_only,
        "use_columns": use_columns,
        "preview_rows": preview_rows,
        "row_count": int(dataframe.height),
        "column_count": int(dataframe.width),
        "columns": list(dataframe.columns),
        "schema": {key: str(value) for key, value in dataframe.schema.items()},
        "rows_preview": rows_preview,
        "size_bytes": resolved.stat().st_size,
    }

